"""OTTER per-disorder spatial pattern prediction (Phase 1 of ENIGMA validation).

Generate per-parcel predicted human spatial patterns for each of 5 psychiatric
disorders + autism, using OTTER's π applied to each disorder's gene set:

  - Autism (Pagani MOESM4 'subtypes', hypo + hyper combined)
  - Bipolar disorder (MOESM5)
  - Schizophrenia (MOESM5)
  - ADHD (MOESM5)
  - Dementia (MOESM5)
  - Psoriasis (MOESM5, non-brain control)

For each disorder, intersect its gene list with OTTER's 1,713-gene Allen ISH
panel, compute the mouse-parcel mean expression score, route through π →
predicted human-parcel score (2094-vec). Save these as `predicted_pattern.npy`
for Phase 2 comparison against ENIGMA observed disease spatial maps.

The script also computes the pairwise cross-disorder correlation matrix at
parcel level, which measures how disorder-specific OTTER's predictions are
independently of ENIGMA. If all disorders predict similar patterns, that
matches the shared-geometry result of the per-network specificity test
(cross_disease_specificity.py).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from scipy.stats import pearsonr, spearmanr

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "src"))

from otter.data import load_cached, load_pi, pi_provenance

MOESM4 = ROOT / "data_external" / "pagani_2026" / "41593_2026_2287_MOESM4_ESM.xlsx"
MOESM5 = ROOT / "data_external" / "pagani_2026" / "41593_2026_2287_MOESM5_ESM.xlsx"


def load_disorder_gene_sets() -> dict[str, set[str]]:
    """Combine autism (MOESM4) + 5 other conditions (MOESM5) into one dict."""
    out: dict[str, set[str]] = {}
    # Autism: union of hypo + hyper
    wb = openpyxl.load_workbook(MOESM4, data_only=True)
    ws = wb["subtypes"]
    autism = set()
    for r in range(2, ws.max_row + 1):
        for col in [1, 2]:
            v = ws.cell(r, col).value
            if v: autism.add(str(v).strip().lower())
    out["autism"] = {g[0].upper() + g[1:].lower() if g else g for g in autism}
    # Other 5 from MOESM5
    wb = openpyxl.load_workbook(MOESM5, data_only=True)
    ws = wb["other_conditions"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for c, h in enumerate(headers, start=1):
        if not h: continue
        cond_genes = set()
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, c).value
            if v:
                s = str(v).strip()
                if s: cond_genes.add(s[0].upper() + s[1:].lower())
        out[h] = cond_genes
    return out


def main():
    print("=" * 80)
    print("OTTER per-disorder spatial pattern prediction (Phase 1)")
    print("=" * 80)

    # ---- Load OTTER expanded gene matrix ----
    expr = np.load(ROOT / "experiments/autism_subtypes/allen_expansion/pagani_mouse_expr.npy")
    meta = pd.read_csv(ROOT / "experiments/autism_subtypes/allen_expansion/pagani_gene_list_resolved.csv")
    pi = load_pi()
    prov = pi_provenance()
    H, _ = load_cached("human", cache_dir=str(ROOT / "outputs/anndata"))
    print(f"OTTER expanded matrix: {expr.shape}")
    print(f"π: {pi.shape}  {prov['pi_file']}  sha256={prov['pi_sha256']}")

    # NaN-fill + z-score
    expr = expr.copy()
    cmean = np.nanmean(expr, axis=0)
    nz = np.where(np.isnan(expr))
    expr[nz] = np.take(cmean, nz[1])
    z = (expr - expr.mean(0, keepdims=True)) / (expr.std(0, keepdims=True) + 1e-9)

    gene_to_idx = {g.lower(): i for i, g in enumerate(meta["mouse_symbol"])}

    # ---- Load disorder gene sets ----
    print("\nLoading disorder gene sets from MOESM4 + MOESM5...")
    disorders = load_disorder_gene_sets()
    for d, gs in disorders.items():
        n_overlap = len({g.lower() for g in gs} & set(gene_to_idx.keys()))
        print(f"  {d:<22s}: {len(gs):>5d} genes, {n_overlap} overlap with OTTER 1,713-gene panel")

    # ---- For each disorder, route through π → predicted human pattern ----
    print(f"\n{'='*80}")
    print(f"Generating predicted human spatial patterns per disorder")
    print(f"{'='*80}")
    predicted = {}
    n_overlap = {}
    for disorder, gene_set in disorders.items():
        cond_lower = {g.lower() for g in gene_set}
        cond_idx = [gene_to_idx[g] for g in cond_lower if g in gene_to_idx]
        n_overlap[disorder] = len(cond_idx)
        if len(cond_idx) < 10:
            print(f"  {disorder:<22s}: {len(cond_idx)} overlap genes, skipped (insufficient)")
            continue
        # Mouse-parcel score = mean of z-scored marker genes
        mouse_score = z[:, cond_idx].mean(axis=1)
        # Translate through π
        pred_human = mouse_score @ pi   # (2094,)
        predicted[disorder] = pred_human
        print(f"  {disorder:<22s}: predicted human pattern computed from {len(cond_idx)} genes "
              f"(range [{pred_human.min():+.3f}, {pred_human.max():+.3f}])")

    # ---- Cross-disorder correlation matrix at parcel level ----
    print(f"\n{'='*80}")
    print(f"Cross-disorder Pearson correlation matrix (2,094 parcels)")
    print(f"{'='*80}")
    valid = list(predicted.keys())
    n = len(valid)
    corr_mat = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            r, _ = pearsonr(predicted[valid[i]], predicted[valid[j]])
            corr_mat[i, j] = r
    print(f"\n{'':<22s}" + "".join(f"{d[:8]:>10s}" for d in valid))
    for i, d in enumerate(valid):
        row = "".join(f"{corr_mat[i,j]:>+10.3f}" for j in range(n))
        print(f"  {d:<20s}{row}")

    # Off-diagonal statistics
    iu = np.triu_indices(n, k=1)
    off_diag = corr_mat[iu]
    print(f"\nOff-diagonal correlation: mean = {off_diag.mean():+.3f}, "
          f"min = {off_diag.min():+.3f}, max = {off_diag.max():+.3f}")
    print(f"  If OTTER predictions are disorder-specific: low off-diagonal (close to 0).")
    print(f"  If OTTER predictions reflect shared geometry: high off-diagonal (close to 1).")

    # ---- Save ----
    np.savez(ROOT / "outputs/coupling/per_disorder_predictions.npz",
             **{f"{d}": predicted[d] for d in valid})
    out = {
        **prov,
        "n_disorders":      len(valid),
        "disorders":        valid,
        "n_gene_overlap":   {d: int(n) for d, n in n_overlap.items()},
        "correlation_matrix": corr_mat.tolist(),
        "off_diag_mean":    float(off_diag.mean()),
        "off_diag_min":     float(off_diag.min()),
        "off_diag_max":     float(off_diag.max()),
    }
    out_path = ROOT / "outputs" / "logs" / "enigma_phase1_per_disorder.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, indent=2))
    print(f"\nWrote {out_path}")
    print(f"Wrote {ROOT / 'outputs/coupling/per_disorder_predictions.npz'}")


if __name__ == "__main__":
    main()
