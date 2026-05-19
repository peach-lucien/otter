"""Test 3 — Gene-set spatial-pattern translation through π (proof of concept).

Pagani 2026 claim 4: the gene/pathway signature of each FC subtype recurs
cross-species. They report this as parallel observations (mouse-subtype gene set
is enriched for synaptic pathways; human-subtype gene set is enriched for the
same pathways) — without explicitly linking the *spatial* expression patterns
through a cross-species mapping. HOMER's π lets us do that explicit linking:
translate the mouse spatial expression map of the hypo/hyper gene sets through
π and check whether the predicted human map aligns with the observed human
subtype perturbation map.

Limitation: HOMER's curated mouse expression matrix contains only 51 genes;
only 36 of these (10 hypo + 26 hyper) overlap with Pagani's gene lists. So
this is a **proof-of-concept**, not a full pathway-spatial test. A full version
would need parcel-level expression for all 1,952 + 4,463 genes from the Allen
Brain Atlas (~3-4 hours of API queries beyond what HOMER ships).

Procedure:
  1. Identify HOMER genes that overlap with Pagani's hypo-only / hyper-only sets.
  2. Compute per-parcel mouse expression score = mean of those gene columns.
  3. Translate via π: pred_human[h] = Σ_m π[m, h] · score[m].
  4. Aggregate to 8 human networks (Pagani's scheme).
  5. Predict Δ_human = (hyper_pred − hypo_pred) per network.
  6. Compare to observed human Δ from Fig 4e (correlation; permuted-π null).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from scipy.stats import pearsonr, spearmanr

sys.path.insert(0, str(Path(__file__).parent))
from importlib import import_module
nc = import_module("01_network_crossvalidation")
st = import_module("04_subtype_translation")
fm = import_module("07_full_matrix_translation")
assign_human_paper_networks = nc.assign_human_paper_networks
assign_mouse_pagani_networks = fm.assign_mouse_pagani_networks
load_pagani_subtype_matrices = st.load_pagani_subtype_matrices
network_intensity = st.network_intensity

from homer.data import load_cached


PAGANI_GENES_PATH = "/sessions/wizardly-admiring-tesla/mnt/uploads/41593_2026_2287_MOESM4_ESM.xlsx"


def load_pagani_gene_sets() -> tuple[set, set]:
    """Return (hypo_only, hyper_only) gene-name sets (lowercased)."""
    wb = openpyxl.load_workbook(PAGANI_GENES_PATH, data_only=True)
    ws = wb["subtypes"]
    hypo = set(); hyper = set()
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a: hypo.add(str(a).strip().lower())
        if b: hyper.add(str(b).strip().lower())
    return hypo - hyper, hyper - hypo


def main():
    print("=" * 80)
    print("Pagani 2026 Test 3 — gene-set spatial translation through π")
    print("=" * 80)

    M, _ = load_cached("mouse", cache_dir="outputs/anndata")
    H, _ = load_cached("human", cache_dir="outputs/anndata")
    pi = np.load("outputs/coupling/pi_fc_plus_SC_with_all_packs.npy")
    # mouse_genes.npy is the unfiltered (1864, 61) matrix matching mouse_gene_list.csv
    mouse_expr = np.load("data_external/mouse_genes.npy")
    homer_genes = pd.read_csv("data_external/mouse_gene_list.csv")["gene_symbol"].tolist()
    homer_lower = [g.lower() for g in homer_genes]
    # Replace NaNs with column means
    col_mean = np.nanmean(mouse_expr, axis=0)
    inds = np.where(np.isnan(mouse_expr))
    mouse_expr = mouse_expr.copy()
    mouse_expr[inds] = np.take(col_mean, inds[1])
    print(f"\nHOMER gene matrix: {mouse_expr.shape}  ({len(homer_genes)} genes)")

    hypo_set, hyper_set = load_pagani_gene_sets()
    print(f"Pagani hypo-only: {len(hypo_set)}, hyper-only: {len(hyper_set)}")

    hypo_idx = [i for i, g in enumerate(homer_lower) if g in hypo_set]
    hyper_idx = [i for i, g in enumerate(homer_lower) if g in hyper_set]
    print(f"\nGene overlap:")
    print(f"  hypo:  {len(hypo_idx)} HOMER genes - {[homer_genes[i] for i in hypo_idx]}")
    print(f"  hyper: {len(hyper_idx)} HOMER genes - {[homer_genes[i] for i in hyper_idx]}")

    if not hypo_idx or not hyper_idx:
        print("Insufficient overlap, aborting.")
        return

    # Per-parcel mouse expression score for each gene set
    # Z-score genes first to make them comparable across different absolute scales
    mE = (mouse_expr - mouse_expr.mean(0, keepdims=True)) / (mouse_expr.std(0, keepdims=True) + 1e-9)
    hypo_score_mouse  = mE[:, hypo_idx].mean(axis=1)    # (1864,)
    hyper_score_mouse = mE[:, hyper_idx].mean(axis=1)   # (1864,)
    print(f"\nMouse-side scores per parcel:")
    print(f"  hypo  range: {hypo_score_mouse.min():+.3f} .. {hypo_score_mouse.max():+.3f}")
    print(f"  hyper range: {hyper_score_mouse.min():+.3f} .. {hyper_score_mouse.max():+.3f}")

    # Translate to human via π
    hypo_pred_human  = hypo_score_mouse  @ pi   # (2094,)
    hyper_pred_human = hyper_score_mouse @ pi   # (2094,)

    # Aggregate predicted human-parcel scores to 8 Pagani networks
    human_net, human_paper_names = assign_human_paper_networks(H.var, separate_aud=True)
    aud_idx = human_paper_names.index("Auditory")
    som_idx = human_paper_names.index("SomatoMotor")
    human_net = human_net.copy()
    human_net[human_net == aud_idx] = som_idx
    pagani_human_names = ["Control", "DMN", "DorsAtten", "Limbic",
                          "Salience", "SomatoMotor", "Visual", "Subcortical"]
    h_name_to_idx = {n: human_paper_names.index(n) for n in pagani_human_names}
    pagani_human_net = np.full_like(human_net, -1)
    for new_i, n in enumerate(pagani_human_names):
        pagani_human_net[human_net == h_name_to_idx[n]] = new_i

    def aggregate_to_pagani(values):
        out = np.zeros(len(pagani_human_names))
        for i in range(len(pagani_human_names)):
            mask = pagani_human_net == i
            if mask.any():
                out[i] = values[mask].mean()
        return out

    hypo_human_8  = aggregate_to_pagani(hypo_pred_human)
    hyper_human_8 = aggregate_to_pagani(hyper_pred_human)

    # Predicted Δ = hyper-spatial - hypo-spatial
    pred_delta = hyper_human_8 - hypo_human_8

    # Observed Δ from Fig 4e
    data = load_pagani_subtype_matrices()
    obs_hypo_int  = network_intensity(data["human_hypo"],  "abs_rowcol_sum")
    obs_hyper_int = network_intensity(data["human_hyper"], "abs_rowcol_sum")
    obs_delta = obs_hyper_int - obs_hypo_int

    print(f"\nPer-network Δ (hyper − hypo) by 8 Pagani human networks:")
    print(f"  {'network':18s} | {'observed':>10s} | {'pred (gene)':>12s} | {'pred z':>8s} | {'obs z':>7s}")
    pred_z = (pred_delta - pred_delta.mean()) / (pred_delta.std() + 1e-9)
    obs_z = (obs_delta - obs_delta.mean()) / (obs_delta.std() + 1e-9)
    for i, n in enumerate(pagani_human_names):
        print(f"  {n:18s} | {obs_delta[i]:>+10.2f} | {pred_delta[i]:>+12.4f} | "
              f"{pred_z[i]:>+8.2f} | {obs_z[i]:>+7.2f}")

    r_p, p_p = pearsonr(pred_delta, obs_delta)
    r_s, p_s = spearmanr(pred_delta, obs_delta)
    same_sign = int((np.sign(pred_delta) == np.sign(obs_delta)).sum())
    print(f"\nPredicted (from gene-spatial) vs Observed (Pagani FC Δ) correlation:")
    print(f"  Pearson  r = {r_p:+.3f} (p = {p_p:.3f})")
    print(f"  Spearman ρ = {r_s:+.3f} (p = {p_s:.3f})")
    print(f"  Same-sign per network: {same_sign}/8")

    # Permuted-π null
    rng = np.random.default_rng(seed=42)
    n_trials = 200
    null_p, null_s = [], []
    for _ in range(n_trials):
        perm = rng.permutation(pi.shape[0])
        pi_n = pi[perm]
        hypo_n  = aggregate_to_pagani(hypo_score_mouse  @ pi_n)
        hyper_n = aggregate_to_pagani(hyper_score_mouse @ pi_n)
        pred_n = hyper_n - hypo_n
        rpn, _ = pearsonr(pred_n, obs_delta)
        rsn, _ = spearmanr(pred_n, obs_delta)
        null_p.append(rpn); null_s.append(rsn)
    null_p = np.array(null_p); null_s = np.array(null_s)
    emp_p_p = float(np.mean(null_p >= r_p))
    emp_p_s = float(np.mean(null_s >= r_s))
    print(f"\nPermuted-π null (200 trials):")
    print(f"  Pearson  null mean={null_p.mean():+.3f}, "
          f"95% CI ({np.percentile(null_p, 2.5):+.3f}, {np.percentile(null_p, 97.5):+.3f}), "
          f"empirical p={emp_p_p:.3f}")
    print(f"  Spearman null mean={null_s.mean():+.3f}, "
          f"95% CI ({np.percentile(null_s, 2.5):+.3f}, {np.percentile(null_s, 97.5):+.3f}), "
          f"empirical p={emp_p_s:.3f}")

    out = {
        "n_homer_genes": len(homer_genes),
        "n_overlap_hypo_genes":  len(hypo_idx),
        "n_overlap_hyper_genes": len(hyper_idx),
        "homer_hypo_genes":  [homer_genes[i] for i in hypo_idx],
        "homer_hyper_genes": [homer_genes[i] for i in hyper_idx],
        "pagani_human_networks": pagani_human_names,
        "observed_delta":  obs_delta.tolist(),
        "predicted_delta": pred_delta.tolist(),
        "same_sign_count": same_sign,
        "pearson_r": float(r_p), "pearson_p": float(p_p),
        "spearman_r": float(r_s), "spearman_p": float(p_s),
        "null": {
            "n_trials": n_trials,
            "pearson_mean":  float(null_p.mean()),
            "pearson_ci95":  [float(np.percentile(null_p, 2.5)), float(np.percentile(null_p, 97.5))],
            "spearman_mean": float(null_s.mean()),
            "spearman_ci95": [float(np.percentile(null_s, 2.5)), float(np.percentile(null_s, 97.5))],
            "pearson_empirical_p":  emp_p_p,
            "spearman_empirical_p": emp_p_s,
        },
        "caveat": ("HOMER's curated gene set is 51 genes; only 36 overlap with Pagani's "
                   "1,952+4,463 implicated genes. This is a proof-of-concept; a full "
                   "pathway-spatial test would need parcel-level Allen Brain Atlas "
                   "expression for the full gene lists."),
    }
    out_path = Path("outputs/logs/autism_subtypes_gene_spatial.json")
    out_path.write_text(json.dumps(out, indent=2))
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
