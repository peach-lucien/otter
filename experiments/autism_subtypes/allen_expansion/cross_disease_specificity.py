"""Cross-disease specificity test for OTTER's gene-spatial cross-species signal.

Test 3 showed that translating Pagani's 1,713 autism-implicated genes through
OTTER's π reliably (bootstrap r=+0.428, 95% CI +0.349–+0.497) predicts Pagani's
observed human ASD subtype Δ pattern.

This script tests whether the signal is autism-specific or whether OTTER
produces the same correlation for any brain-disorder gene set. Pagani's MOESM5
supplementary lists 4,822 genes implicated in five comparison conditions:
  - bipolar_disorder
  - schizophrenia
  - psoriasis      (non-brain control, strong negative control)
  - dementia
  - adhd

For each condition:
  1. Take its gene list.
  2. Intersect with OTTER's 1,713-gene Allen ISH matrix.
  3. Compute mean z-scored expression per mouse parcel.
  4. Translate via π to predicted per-human-parcel score.
  5. Aggregate to 8 Pagani networks.
  6. Correlate against Pagani's observed human ASD Δ (hyper−hypo, Fig 4e).
  7. Bootstrap (200 resamples of the gene pool) to get CI per condition.

Under autism-specificity:
  - Autism: r ≈ +0.43 (replicates Test 3)
  - Brain psychiatric conditions (bipolar, schizophrenia, adhd, dementia):
    moderate correlations, since they share neural gene biology with autism
  - Psoriasis: r ≈ 0 (skin disease, no brain relevance, negative control)

If autism's r is similar to schizophrenia's, the signal is not autism-specific.
If psoriasis also produces r ≈ +0.4, the signal is a brain-geometry artefact
rather than a disease-specific finding.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from scipy.stats import pearsonr, spearmanr

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "experiments" / "autism_subtypes"))

from importlib import import_module
nc = import_module("01_network_crossvalidation")
st = import_module("04_subtype_translation")

from otter.data import load_cached


MOESM5_PATH = ROOT / "data_external" / "pagani_2026" / "41593_2026_2287_MOESM5_ESM.xlsx"


def load_other_disease_genes() -> dict[str, set[str]]:
    """Return dict condition_name → set of mouse-cased gene symbols."""
    wb = openpyxl.load_workbook(MOESM5_PATH, data_only=True)
    ws = wb["other_conditions"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    out: dict[str, set[str]] = {h: set() for h in headers if h}
    for r in range(2, ws.max_row + 1):
        for c, h in enumerate(headers, start=1):
            if not h: continue
            v = ws.cell(r, c).value
            if v:
                s = str(v).strip()
                if s:
                    out[h].add(s[0].upper() + s[1:].lower())
    return out


def load_pagani_subtype_gene_sets():
    """Reload autism subtype gene sets from MOESM4 for the comparison."""
    from openpyxl import load_workbook
    p = ROOT / "data_external" / "pagani_2026" / "41593_2026_2287_MOESM4_ESM.xlsx"
    wb = load_workbook(p, data_only=True)
    ws = wb["subtypes"]
    hypo, hyper = set(), set()
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a: hypo.add(str(a).strip().lower())
        if b: hyper.add(str(b).strip().lower())
    return hypo, hyper


def main():
    print("=" * 80)
    print("Cross-disease specificity test")
    print("=" * 80)

    # Load OTTER expanded matrix + metadata
    expr = np.load(ROOT / "experiments/autism_subtypes/allen_expansion/pagani_mouse_expr.npy")
    meta = pd.read_csv(ROOT / "experiments/autism_subtypes/allen_expansion/pagani_gene_list_resolved.csv")
    pi = np.load(ROOT / "outputs/coupling/pi_fc_plus_SC_with_all_packs.npy")
    H, _ = load_cached("human", cache_dir=str(ROOT / "outputs" / "anndata"))
    print(f"\nOTTER expanded matrix: {expr.shape}")

    # NaN-fill + z-score
    expr = expr.copy()
    cmean = np.nanmean(expr, axis=0)
    nz = np.where(np.isnan(expr))
    expr[nz] = np.take(cmean, nz[1])
    z = (expr - expr.mean(0, keepdims=True)) / (expr.std(0, keepdims=True) + 1e-9)

    # OTTER gene index by lowercased symbol
    gene_to_idx = {g.lower(): i for i, g in enumerate(meta["mouse_symbol"])}

    # Disease gene sets
    diseases = load_other_disease_genes()
    print(f"\nLoaded comparison conditions from MOESM5: {list(diseases.keys())}")
    for d, gs in diseases.items():
        print(f"  {d:25s}: {len(gs):4d} genes")

    # Add autism (combined hypo + hyper) as one condition
    autism_hypo, autism_hyper = load_pagani_subtype_gene_sets()
    autism_combined = autism_hypo | autism_hyper
    print(f"  autism_combined (hypo+hyper): {len(autism_combined)} genes")
    diseases["autism_combined"] = {g[0].upper() + g[1:].lower() if g else g for g in autism_combined}

    # Observed human ASD Δ from Fig 4e
    data = st.load_pagani_subtype_matrices()
    obs_hypo  = st.network_intensity(data["human_hypo"], "abs_rowcol_sum")
    obs_hyper = st.network_intensity(data["human_hyper"], "abs_rowcol_sum")
    obs_delta = obs_hyper - obs_hypo

    # Human network assignment (Pagani's 8 networks)
    human_net, human_paper_names = nc.assign_human_paper_networks(H.var, separate_aud=True)
    aud_idx = human_paper_names.index("Auditory")
    som_idx = human_paper_names.index("SomatoMotor")
    human_net = human_net.copy()
    human_net[human_net == aud_idx] = som_idx
    pagani_human = ["Control", "DMN", "DorsAtten", "Limbic", "Salience",
                    "SomatoMotor", "Visual", "Subcortical"]
    h_name_to_idx = {n: human_paper_names.index(n) for n in pagani_human}
    pag_net = np.full_like(human_net, -1)
    for new_i, n in enumerate(pagani_human):
        pag_net[human_net == h_name_to_idx[n]] = new_i

    def agg8(values):
        out = np.zeros(len(pagani_human))
        for i in range(len(pagani_human)):
            mask = pag_net == i
            if mask.any():
                out[i] = values[mask].mean()
        return out

    # Run the test per condition
    print(f"\n{'Condition':<22s} | {'overlap':>7s} | {'r_obs_delta':>12s} | "
          f"{'bootstrap mean (95% CI)':>30s}")
    print("-" * 95)
    rng = np.random.default_rng(seed=42)
    n_boot = 500
    results = {}
    for cond, gene_set in diseases.items():
        cond_lower = {g.lower() for g in gene_set}
        cond_idx = [gene_to_idx[g] for g in cond_lower if g in gene_to_idx]
        if len(cond_idx) < 10:
            print(f"  {cond:<22s} | {len(cond_idx):>7d} | (too few overlapping genes, skipped)")
            results[cond] = {"n_overlap": len(cond_idx), "skipped": True}
            continue
        # Per-parcel score
        score_mouse = z[:, cond_idx].mean(axis=1)
        # Translate to human + aggregate to 8 networks
        pred_h = agg8(score_mouse @ pi)
        # Subtract the across-network mean so the predicted is comparable to Δ (signed)
        # (the observed uses abs_rowcol_sum but pred is per-network expression intensity;
        # both are means plus a constant, so Pearson is invariant and no subtraction
        # is applied.)
        r_obs, p_obs = pearsonr(pred_h, obs_delta)
        # Bootstrap over genes
        boot_r = []
        for _ in range(n_boot):
            samp = rng.choice(cond_idx, size=len(cond_idx), replace=True)
            sm = z[:, samp].mean(axis=1)
            ph = agg8(sm @ pi)
            r, _ = pearsonr(ph, obs_delta)
            boot_r.append(r)
        boot_r = np.array(boot_r)
        results[cond] = {
            "n_overlap":   int(len(cond_idx)),
            "r_observed":  float(r_obs),
            "p_analytical": float(p_obs),
            "boot_mean":   float(boot_r.mean()),
            "boot_ci95":   [float(np.percentile(boot_r, 2.5)),
                            float(np.percentile(boot_r, 97.5))],
            "pct_positive": float((boot_r > 0).mean()),
        }
        print(f"  {cond:<22s} | {len(cond_idx):>7d} | {r_obs:>+12.3f} | "
              f"{boot_r.mean():>+7.3f} ({np.percentile(boot_r, 2.5):+.3f}, "
              f"{np.percentile(boot_r, 97.5):+.3f})  | {(boot_r>0).mean()*100:.0f}% positive")

    out_path = ROOT / "outputs" / "logs" / "autism_subtypes_cross_disease.json"
    out_path.write_text(json.dumps({"n_bootstraps": n_boot, "results": results}, indent=2))
    print(f"\nWrote {out_path}")

    # Rank-order the conditions and write a small interpretation
    valid = [(c, r) for c, r in results.items() if not r.get("skipped")]
    valid.sort(key=lambda x: -x[1]["boot_mean"])
    print(f"\nRanked by bootstrap mean r (high → low):")
    for c, r in valid:
        print(f"  {c:<22s}: {r['boot_mean']:+.3f}  (95% CI {r['boot_ci95'][0]:+.3f}..{r['boot_ci95'][1]:+.3f})")

    # Verdict
    autism_r = results.get("autism_combined", {}).get("boot_mean", 0)
    psoriasis_r = results.get("psoriasis", {}).get("boot_mean", 0)
    if autism_r and psoriasis_r:
        ratio = autism_r / max(abs(psoriasis_r), 0.01)
        print(f"\nAutism vs psoriasis (non-brain control): "
              f"{autism_r:+.3f} vs {psoriasis_r:+.3f}  (ratio: {ratio:.1f}×)")
        if abs(psoriasis_r) > 0.2 and abs(psoriasis_r) > 0.5 * abs(autism_r):
            print("  → Signal NOT autism-specific (psoriasis comparable to autism)")
        elif abs(psoriasis_r) < 0.15 and autism_r > 0.3:
            print("  → Signal IS autism-specific (psoriasis near zero; autism strong)")
        else:
            print("  → Signal partially autism-specific (psoriasis modest, autism stronger)")


if __name__ == "__main__":
    main()
