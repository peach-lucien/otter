"""Diagnostic + sharper analysis of the expanded Pagani gene matrix.

Three checks:
  1. Per-parcel correlation between predicted-from-gene-spatial and
     predicted-from-FC (Test 2c style). Does the gene translation agree with
     the FC translation at parcel level? n=2,094 instead of 8.
  2. Gene-bootstrap: subsample with replacement to estimate uncertainty on
     the predicted Δ pattern.
  3. Pathway-by-pathway test: instead of one hypo + one hyper bulk score,
     test each pathway separately using MOESM3's per-pathway gene lists.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from scipy.stats import pearsonr, spearmanr

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "experiments" / "autism_subtypes"))

from importlib import import_module
nc = import_module("01_network_crossvalidation")
st = import_module("04_subtype_translation")
fm = import_module("07_full_matrix_translation")

from homer.data import load_cached


PAGANI_PATHWAYS_PATH = ROOT.parent.parent.parent.parent / "uploads" / "41593_2026_2287_MOESM3_ESM.xlsx"
# Fallback: try the canonical session uploads path
if not PAGANI_PATHWAYS_PATH.exists():
    PAGANI_PATHWAYS_PATH = Path("/sessions/wizardly-admiring-tesla/mnt/uploads/41593_2026_2287_MOESM3_ESM.xlsx")


def load_pathway_genes() -> dict[str, set]:
    """Return dict pathway_name → set of mouse-symbol-cased gene names."""
    wb = openpyxl.load_workbook(PAGANI_PATHWAYS_PATH, data_only=True)
    ws = wb["pathways"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    out: dict[str, set] = {h: set() for h in headers if h}
    for r in range(2, ws.max_row+1):
        for c, h in enumerate(headers, start=1):
            if not h: continue
            v = ws.cell(r, c).value
            if v:
                s = str(v).strip()
                if s: out[h].add(s[0].upper() + s[1:].lower())
    return out


def main():
    print("=" * 80)
    print("Pagani 2026 — diagnostic checks on the expanded gene matrix")
    print("=" * 80)

    expr = np.load(ROOT / "experiments/autism_subtypes/allen_expansion/pagani_mouse_expr.npy")
    meta = pd.read_csv(ROOT / "experiments/autism_subtypes/allen_expansion/pagani_gene_list_resolved.csv")
    pi = np.load(ROOT / "outputs/coupling/pi_fc_plus_SC_with_all_packs.npy")
    H, _ = load_cached("human", cache_dir=str(ROOT / "outputs" / "anndata"))
    M, _ = load_cached("mouse", cache_dir=str(ROOT / "outputs" / "anndata"))

    # NaN handling + z-score
    expr = expr.copy()
    cmean = np.nanmean(expr, axis=0)
    nz = np.where(np.isnan(expr))
    expr[nz] = np.take(cmean, nz[1])
    z = (expr - expr.mean(0, keepdims=True)) / (expr.std(0, keepdims=True) + 1e-9)

    hypo_idx = np.where(meta["subtype"] == "hypo")[0]
    hyper_idx = np.where(meta["subtype"] == "hyper")[0]
    hypo_mouse_pp  = z[:, hypo_idx].mean(axis=1)
    hyper_mouse_pp = z[:, hyper_idx].mean(axis=1)

    # Translate to human via pi (per-parcel)
    hypo_human_pp  = hypo_mouse_pp  @ pi
    hyper_human_pp = hyper_mouse_pp @ pi
    pred_delta_pp_gene = hyper_human_pp - hypo_human_pp   # 2,094-vec

    # ---- DIAGNOSTIC 1: per-parcel correlation between predicted-from-gene and
    # predicted-from-FC -------------------------------------------------------
    print("\n--- Diagnostic 1: per-parcel agreement between gene-translation and "
          "FC-translation (n=2,094) ---")
    # Build predicted-from-FC per-parcel Δ exactly like Test 2c does (mouse 9×9 →
    # parcel-level → π).
    data = st.load_pagani_subtype_matrices()
    mouse_pagani_net, mouse_pagani_names = fm.assign_mouse_pagani_networks(M.var)
    keep = mouse_pagani_net >= 0

    def _intensity(MM):
        Ma = np.abs(MM)
        return Ma.sum(axis=0) + Ma.sum(axis=1) - np.diag(Ma)

    hypo_net  = _intensity(0.5 * (data["mouse_hypo"]  + data["mouse_hypo"].T))
    hyper_net = _intensity(0.5 * (data["mouse_hyper"] + data["mouse_hyper"].T))

    def _to_mouse_parcel(intensity, mouse_paper_net, kept_mask):
        v = np.zeros(pi.shape[0])
        for i in range(len(mouse_pagani_names)):
            v[(mouse_paper_net == i) & kept_mask] = intensity[i]
        return v

    hypo_fc_parcel  = _to_mouse_parcel(hypo_net,  mouse_pagani_net, keep)
    hyper_fc_parcel = _to_mouse_parcel(hyper_net, mouse_pagani_net, keep)
    pred_delta_pp_FC = (hyper_fc_parcel - hypo_fc_parcel) @ pi   # 2,094-vec

    r_gene_vs_fc_p, p_gv = pearsonr(pred_delta_pp_gene, pred_delta_pp_FC)
    r_gene_vs_fc_s, p_gvs = spearmanr(pred_delta_pp_gene, pred_delta_pp_FC)
    print(f"  Per-parcel Pearson r (gene-translation vs FC-translation) = {r_gene_vs_fc_p:+.3f}")
    print(f"  Per-parcel Spearman ρ = {r_gene_vs_fc_s:+.3f}")
    print(f"  Both predictions encoded the same biology? {'YES' if r_gene_vs_fc_p > 0.3 else 'PARTIAL' if r_gene_vs_fc_p > 0.1 else 'NO'}")

    # ---- DIAGNOSTIC 2: bootstrap over genes --------------------------------
    print("\n--- Diagnostic 2: gene bootstrap (1000 resamples of the gene pool) ---")
    # Reload the 8-network observed delta
    obs_hypo  = st.network_intensity(data["human_hypo"],  "abs_rowcol_sum")
    obs_hyper = st.network_intensity(data["human_hyper"], "abs_rowcol_sum")
    obs_delta = obs_hyper - obs_hypo

    human_net, human_paper_names = nc.assign_human_paper_networks(H.var, separate_aud=True)
    aud_idx = human_paper_names.index("Auditory")
    som_idx = human_paper_names.index("SomatoMotor")
    human_net = human_net.copy()
    human_net[human_net == aud_idx] = som_idx
    pagani_human = ["Control", "DMN", "DorsAtten", "Limbic",
                    "Salience", "SomatoMotor", "Visual", "Subcortical"]
    h_name_to_idx = {n: human_paper_names.index(n) for n in pagani_human}
    pag_net = np.full_like(human_net, -1)
    for new_i, n in enumerate(pagani_human):
        pag_net[human_net == h_name_to_idx[n]] = new_i

    def agg8(values):
        out = np.zeros(len(pagani_human))
        for i in range(len(pagani_human)):
            m = pag_net == i
            if m.any():
                out[i] = values[m].mean()
        return out

    rng = np.random.default_rng(42)
    n_boot = 1000
    boot_r = []
    n_hypo, n_hyper = len(hypo_idx), len(hyper_idx)
    for _ in range(n_boot):
        sb_hypo  = rng.choice(hypo_idx,  size=n_hypo,  replace=True)
        sb_hyper = rng.choice(hyper_idx, size=n_hyper, replace=True)
        hp = z[:, sb_hypo].mean(axis=1) @ pi
        hr = z[:, sb_hyper].mean(axis=1) @ pi
        d8 = agg8(hr) - agg8(hp)
        r, _ = pearsonr(d8, obs_delta)
        boot_r.append(r)
    boot_r = np.array(boot_r)
    print(f"  Bootstrap Pearson r: mean = {boot_r.mean():+.3f}, "
          f"95% CI ({np.percentile(boot_r, 2.5):+.3f}, {np.percentile(boot_r, 97.5):+.3f})")
    print(f"  Fraction of bootstraps with r > 0:  {(boot_r > 0).mean()*100:.1f}%")
    print(f"  Fraction of bootstraps with r > 0.3: {(boot_r > 0.3).mean()*100:.1f}%")

    # ---- DIAGNOSTIC 3: pathway-by-pathway --------------------------------
    print("\n--- Diagnostic 3: per-pathway test (each Pagani pathway separately) ---")
    pathways = load_pathway_genes()
    homer_genes = meta["mouse_symbol"].tolist()
    homer_lower = {g.lower(): i for i, g in enumerate(homer_genes)}
    results = []
    for pname, pgenes in pathways.items():
        # which HOMER-resolved genes belong to this pathway?
        pgenes_lower = {g.lower() for g in pgenes}
        gene_indices = [homer_lower[g] for g in homer_lower if g in pgenes_lower]
        if len(gene_indices) < 5:
            continue
        # Pathway-spatial map: mean expression of pathway genes per parcel
        pmap_mouse = z[:, gene_indices].mean(axis=1)
        pmap_human_pred = pmap_mouse @ pi
        pmap_human_8 = agg8(pmap_human_pred)
        # Synaptic pathways should correlate with observed-hypo Δ (negative on delta scale)
        # Immune pathways should correlate with observed-hyper Δ (positive on delta scale)
        r_to_hypo, _ = pearsonr(pmap_human_8, obs_hypo)
        r_to_hyper, _ = pearsonr(pmap_human_8, obs_hyper)
        r_to_delta, _ = pearsonr(pmap_human_8, obs_delta)
        results.append({
            "pathway": pname,
            "n_genes_in_homer": len(gene_indices),
            "r_to_obs_hypo":  float(r_to_hypo),
            "r_to_obs_hyper": float(r_to_hyper),
            "r_to_obs_delta": float(r_to_delta),
        })
    res_df = pd.DataFrame(results).sort_values("r_to_obs_delta")
    print(res_df.to_string(index=False))

    out_json = {
        "n_genes": int(expr.shape[1]),
        "n_hypo": int(len(hypo_idx)),
        "n_hyper": int(len(hyper_idx)),
        "per_parcel_gene_vs_fc_translation": {
            "pearson_r":  float(r_gene_vs_fc_p),
            "spearman_r": float(r_gene_vs_fc_s),
        },
        "bootstrap": {
            "n_resamples": n_boot,
            "mean_r":   float(boot_r.mean()),
            "ci95":     [float(np.percentile(boot_r, 2.5)), float(np.percentile(boot_r, 97.5))],
            "pct_r_positive": float((boot_r > 0).mean()),
            "pct_r_above_0_3": float((boot_r > 0.3).mean()),
        },
        "per_pathway": results,
    }
    out_path = ROOT / "outputs" / "logs" / "autism_subtypes_gene_diagnostics.json"
    out_path.write_text(json.dumps(out_json, indent=2))
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
